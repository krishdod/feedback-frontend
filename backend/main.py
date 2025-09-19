from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import load_workbook, Workbook
from uuid import uuid4
from datetime import datetime
import os
import json
import tempfile

# Google Sheets dependencies
from google.oauth2 import service_account
from googleapiclient.discovery import build

# Environment variables are expected to be provided by the hosting platform
# (e.g., Render) or your local shell. No local setup script is imported.

app = FastAPI()
print("🚀 FastAPI is initializing...")

@app.get("/health")
async def health_check():
    return {"status": "healthy", "message": "Backend is running"}


# Allow all origins for now (customize later)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = "Training Feedback.xlsx"  # Legacy local Excel (kept for compatibility)

# Environment configuration for Google Sheets
GOOGLE_SERVICE_ACCOUNT_JSON = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "")
GOOGLE_SHEETS_SPREADSHEET_ID = os.getenv("GOOGLE_SHEETS_SPREADSHEET_ID", "1BYCStdQILVx8DHFxeRWe5jlISGxgSWikHyuUy_NgoMY")
GOOGLE_SHEETS_TAB_NAME = os.getenv("GOOGLE_SHEETS_TAB_NAME", "Sheet1")

# If no environment variable, try to read from service-account.json file
if not GOOGLE_SERVICE_ACCOUNT_JSON and os.path.exists("service-account.json"):
    with open("service-account.json", "r") as f:
        GOOGLE_SERVICE_ACCOUNT_JSON = f.read()


def get_sheets_service():
    """Create and return a Google Sheets API service using a service account.

    Expects `GOOGLE_SERVICE_ACCOUNT_JSON` env var containing the full JSON
    credentials (either as raw JSON or base64-encoded JSON).
    """
    if not GOOGLE_SERVICE_ACCOUNT_JSON or not GOOGLE_SHEETS_SPREADSHEET_ID:
        return None

    try:
        # Accept raw JSON or base64-encoded
        raw = GOOGLE_SERVICE_ACCOUNT_JSON
        try:
            # Attempt base64 decode; if it fails, treat as raw JSON
            import base64
            decoded = base64.b64decode(raw).decode("utf-8")
            raw = decoded
        except Exception:
            pass

        service_account_info = json.loads(raw)
        credentials = service_account.Credentials.from_service_account_info(
            service_account_info,
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive.file",
            ],
        )
        return build("sheets", "v4", credentials=credentials).spreadsheets()
    except Exception as e:
        print(f"Failed to initialize Google Sheets service: {e}")
        return None


def sheets_get_all_rows():
    service = get_sheets_service()
    if not service:
        return None, None
    range_name = f"{GOOGLE_SHEETS_TAB_NAME}!A:Z"
    result = service.values().get(
        spreadsheetId=GOOGLE_SHEETS_SPREADSHEET_ID,
        range=range_name,
    ).execute()
    values = result.get("values", [])
    if not values:
        return [], []
    headers = values[0]
    rows = values[1:] if len(values) > 1 else []
    return headers, rows


def sheets_append_row(row_values: list[str]):
    service = get_sheets_service()
    if not service:
        return False, "Google Sheets not configured"
    try:
        range_name = f"{GOOGLE_SHEETS_TAB_NAME}!A:Z"
        body = {"values": [row_values]}
        service.values().append(
            spreadsheetId=GOOGLE_SHEETS_SPREADSHEET_ID,
            range=range_name,
            valueInputOption="USER_ENTERED",
            insertDataOption="INSERT_ROWS",
            body=body,
        ).execute()
        return True, None
    except Exception as e:
        print(f"Google Sheets append failed: {e}")
        return False, f"Google Sheets error: {str(e)}"


def sheets_delete_row_by_submission_id(submission_id: str):
    service = get_sheets_service()
    if not service:
        return False, "Google Sheets not configured"

    # Read all rows to find row index (1-based including header)
    headers, rows = sheets_get_all_rows()
    if headers is None:
        return False, "Unable to read Google Sheet"

    # submission_id is in column B per our schema (index 1)
    row_index = None
    for idx, row in enumerate(rows, start=2):  # start=2 because row 1 is headers
        if len(row) > 1 and row[1] == submission_id:
            row_index = idx
            break

    if row_index is None:
        return False, f"Submission with ID {submission_id} not found."

    # Build batchUpdate request to delete the row (0-based indices for requests)
    requests = [
        {
            "deleteDimension": {
                "range": {
                    "sheetId": None,  # Use sheetId if known; fallback to a find-by-title request
                    "dimension": "ROWS",
                    "startIndex": row_index - 1,
                    "endIndex": row_index,
                }
            }
        }
    ]

    # Resolve sheetId by title
    sheets_metadata = build("sheets", "v4", credentials=get_sheets_service()._root._http.credentials).spreadsheets().get(
        spreadsheetId=GOOGLE_SHEETS_SPREADSHEET_ID
    ).execute()
    sheet_id = None
    for s in sheets_metadata.get("sheets", []):
        properties = s.get("properties", {})
        if properties.get("title") == GOOGLE_SHEETS_TAB_NAME:
            sheet_id = properties.get("sheetId")
            break

    if sheet_id is None:
        return False, f"Sheet '{GOOGLE_SHEETS_TAB_NAME}' not found."

    requests[0]["deleteDimension"]["range"]["sheetId"] = sheet_id

    build("sheets", "v4", credentials=get_sheets_service()._root._http.credentials).spreadsheets().batchUpdate(
        spreadsheetId=GOOGLE_SHEETS_SPREADSHEET_ID,
        body={"requests": requests},
    ).execute()

    return True, None


@app.get("/sheets-debug")
async def sheets_debug():
    """Diagnostics for Google Sheets configuration and access.

    Returns a JSON report indicating which step fails: env vars presence,
    credentials decoding, API init, spreadsheet access, tab presence, and a
    small read test. No secrets are returned.
    """
    report = {
        "env": {
            "GOOGLE_SERVICE_ACCOUNT_JSON_present": bool(GOOGLE_SERVICE_ACCOUNT_JSON),
            "GOOGLE_SHEETS_SPREADSHEET_ID_present": bool(GOOGLE_SHEETS_SPREADSHEET_ID),
            "GOOGLE_SHEETS_TAB_NAME": GOOGLE_SHEETS_TAB_NAME,
        },
        "init": None,
        "spreadsheet_access": None,
        "tab_found": None,
        "read_headers": None,
        "errors": [],
    }

    # Try initialize service
    service = None
    try:
        service = get_sheets_service()
        report["init"] = service is not None
        if service is None:
            report["errors"].append("Service init failed (check env vars or JSON key)")
            return report
    except Exception as e:
        report["init"] = False
        report["errors"].append(f"Service init exception: {str(e)}")
        return report

    # Try to fetch spreadsheet metadata
    try:
        meta = build("sheets", "v4", credentials=get_sheets_service()._root._http.credentials).spreadsheets().get(
            spreadsheetId=GOOGLE_SHEETS_SPREADSHEET_ID
        ).execute()
        report["spreadsheet_access"] = {
            "title": meta.get("properties", {}).get("title"),
            "sheets": [s.get("properties", {}).get("title") for s in meta.get("sheets", [])],
        }
        # Check tab exists
        report["tab_found"] = GOOGLE_SHEETS_TAB_NAME in report["spreadsheet_access"]["sheets"]
        if not report["tab_found"]:
            report["errors"].append(f"Tab '{GOOGLE_SHEETS_TAB_NAME}' not found")
    except Exception as e:
        report["spreadsheet_access"] = False
        report["errors"].append(f"Spreadsheet access error: {str(e)}")
        return report

    # Try to read headers
    try:
        headers, rows = sheets_get_all_rows() or (None, None)
        if headers is None:
            report["read_headers"] = False
            report["errors"].append("Read values returned None (likely service not configured)")
        else:
            report["read_headers"] = headers
    except Exception as e:
        report["read_headers"] = False
        report["errors"].append(f"Read error: {str(e)}")

    return report

# Define expected data from frontend
class FeedbackForm(BaseModel):
    full_name: str
    email: str
    job_role: str
    training_title: str
    instructor_name: str
    content_ratings: list[int]
    trainer_ratings: list[int]
    organization_ratings: list[int]
    overall_ratings: list[int]
    covered_topics: list[str]
    other_topic: str
    comments: str

@app.get("/")
async def root():
    return {"message": "Training Feedback API is running!", "endpoints": ["/submit-feedback", "/view-data", "/download-excel", "/docs"]}

@app.get("/health")
async def health_check():
    """Health check endpoint for monitoring and wake-up detection"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "service": "Training Feedback API"
    }

@app.get("/view-data")
async def view_data():
    """View all stored feedback data from Google Sheets if configured; fallback to Excel."""
    try:
        headers, rows = sheets_get_all_rows() or (None, None)
        if headers is not None:
            return {
                "status": "success",
                "total_submissions": len(rows),
                "headers": headers,
                "data": rows,
            }

        # Fallback to legacy Excel file
        if not os.path.exists(EXCEL_FILE):
            return {"status": "error", "message": "No data source configured (Google Sheets or Excel)."}

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        data = []
        headers_excel = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                data.append(row)
        return {
            "status": "success",
            "total_submissions": len(data),
            "headers": headers_excel,
            "data": data,
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/download-excel")
async def download_excel():
    """Generate and download an Excel export from Google Sheets (or Excel fallback)."""
    try:
        headers, rows = sheets_get_all_rows() or (None, None)
        if headers is not None:
            # Normalize to dashboard-friendly headers
            export_headers = [
                "Submitted At",
                "Participant",
                "Email",
                "Role",
                "Training Title",
                "Instructor",
                "Avg Rating",
                "Topics Covered",
                "Comments",
            ]

            def _safe_num(v):
                try:
                    n = float(v)
                    return n
                except Exception:
                    return None

            wb = Workbook()
            ws = wb.active
            ws.append(export_headers)

            for r in rows:
                r = r or []
                submitted = r[0] if len(r) > 0 else ""
                name = r[2] if len(r) > 2 else ""
                email = r[3] if len(r) > 3 else ""
                role = r[4] if len(r) > 4 else ""
                title = r[5] if len(r) > 5 else ""
                instructor = r[6] if len(r) > 6 else ""
                # Prefer overall average at index 10 if present; otherwise compute simple mean of numeric cells
                overall = _safe_num(r[10]) if len(r) > 10 else None
                if overall is None:
                    nums = []
                    for v in r:
                        n = _safe_num(v)
                        if n is not None and n > 0:
                            nums.append(n)
                    overall = round(sum(nums) / len(nums), 1) if nums else ""
                topics = r[11] if len(r) > 11 else ""
                comments = r[13] if len(r) > 13 else ""

                ws.append([
                    submitted,
                    name,
                    email,
                    role,
                    title,
                    instructor,
                    overall,
                    topics,
                    comments,
                ])
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(tmp.name)
            tmp.close()
            return FileResponse(
                path=tmp.name,
                filename="Training_Feedback_Data.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # Fallback to legacy Excel file if present
        if os.path.exists(EXCEL_FILE):
            # Rebuild a normalized export from the legacy workbook
            from openpyxl import load_workbook as _load
            wb_in = _load(EXCEL_FILE)
            ws_in = wb_in.active
            data = list(ws_in.iter_rows(values_only=True))
            legacy_headers = data[0] if data else []
            legacy_rows = data[1:] if len(data) > 1 else []

            export_headers = [
                "Submitted At",
                "Participant",
                "Email",
                "Role",
                "Training Title",
                "Instructor",
                "Avg Rating",
                "Topics Covered",
                "Comments",
            ]
            wb = Workbook()
            ws = wb.active
            ws.append(export_headers)
            for r in legacy_rows:
                r = list(r)
                submitted = r[0] if len(r) > 0 else ""
                name = r[2] if len(r) > 2 else ""
                email = r[3] if len(r) > 3 else ""
                role = r[4] if len(r) > 4 else ""
                title = r[5] if len(r) > 5 else ""
                instructor = r[6] if len(r) > 6 else ""
                overall = r[10] if len(r) > 10 else ""
                topics = r[11] if len(r) > 11 else ""
                comments = r[13] if len(r) > 13 else ""
                ws.append([submitted, name, email, role, title, instructor, overall, topics, comments])

            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(tmp.name)
            tmp.close()
            return FileResponse(
                path=tmp.name,
                filename="Training_Feedback_Data.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        return {"status": "error", "message": "No data available to export."}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.delete("/delete-feedback/{submission_id}")
async def delete_feedback(submission_id: str):
    """Delete a specific feedback entry by submission ID from Google Sheets; fallback to Excel."""
    try:
        ok, err = sheets_delete_row_by_submission_id(submission_id)
        if ok:
            return {"status": "success", "message": f"Submission {submission_id} deleted successfully."}
        elif err != "Google Sheets not configured":
            return {"status": "error", "message": err}

        # Fallback to Excel
        if not os.path.exists(EXCEL_FILE):
            return {"status": "error", "message": "No data source configured (Google Sheets or Excel)."}

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        row_to_delete = None
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[1] == submission_id:
                row_to_delete = row_num
                break
        if row_to_delete is None:
            return {"status": "error", "message": f"Submission with ID {submission_id} not found."}
        ws.delete_rows(row_to_delete)
        wb.save(EXCEL_FILE)
        return {"status": "success", "message": f"Submission {submission_id} deleted successfully."}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/submit-feedback")

async def submit_feedback(form: FeedbackForm):
    try:
        # Prepare row for either Google Sheets or Excel

        submission_id = str(uuid4())[:8]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        def average(lst):
            return round(sum(lst) / len(lst), 2) if lst else 0

        content_avg = average(form.content_ratings)
        trainer_avg = average(form.trainer_ratings)
        org_avg = average(form.organization_ratings)
        overall_avg = average(form.overall_ratings)

        row = [
            timestamp,
            submission_id,
            form.full_name,
            form.email,
            form.job_role,
            form.training_title,
            form.instructor_name,
            content_avg,
            trainer_avg,
            org_avg,
            overall_avg,
            ", ".join(form.covered_topics),
            form.other_topic,
            form.comments,
        ]

        # Try Google Sheets first
        print(f"Attempting to save to Google Sheets...")
        ok, err = sheets_append_row([str(c) if c is not None else "" for c in row])
        if ok:
            print(f"Successfully saved to Google Sheets")
            return {"status": "success", "submission_id": submission_id}
        elif err != "Google Sheets not configured" and "Google Sheets error" not in err:
            print(f"Google Sheets error: {err}")
            return {"status": "error", "message": err}

        # Fallback to legacy Excel storage
        print(f"Falling back to Excel file: {EXCEL_FILE}")
        if not os.path.exists(EXCEL_FILE):
            return {"status": "error", "message": f"{EXCEL_FILE} not found and Google Sheets not configured."}

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append(row)
        wb.save(EXCEL_FILE)
        print(f"Successfully saved to Excel file")

        return {"status": "success", "submission_id": submission_id}
    except Exception as e:
        return {"status": "error", "message": str(e)}

if __name__ == "__main__":
    import uvicorn
    print("✅ FastAPI loaded successfully")
    uvicorn.run(app, host="0.0.0.0", port=9000)
